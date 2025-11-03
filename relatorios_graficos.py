# Biblietacas utilizadas:

import numpy as np
import pandas as pd # Especialista em ler e munipular tabelas.
import matplotlib.pyplot as plt # Utilizado para gerar gráficos.
import io # Salva dados (como imagens de gráficos) na memória.
import os # O 'os' (Sistema Operacional) nos deixa interagir com arquivos e pastas.

from openpyxl import Workbook, drawing  
# É uma ferramenta extra para trabalhar com arquivos Excel (como adicionar gráficos).

arquivo_excel = "bigData.xlsx"
# Define o nome do arquivo Excel que vamos ler.

# Funções para gerar um relatório completo, utilizando o try e o excepect, logo abaixo:
# Observação: As funções de relatórios são praticamentes as mesmas, muda somente na parte de seleção
# de colunas de cada uma e na exibição.
def gerar_relatorio_completo():

    # Caso dê tudo certo ele irá rodar o try:

    # abas = Lê TODAS as abas do Excel de uma vez.
    # relatorio_completo = Junta (Concatena) todas as abas em um único relatório.
    # relatorio final = Limpa o relatório, retirando as partes em branco (NaN).
    # nome_arquivo_saida = Define o nome do novo arquivo Excel.
    # relatorio_final.to_excel() = Salva o relatório limpo nesse novo arquivo).
    # os.startfile() = Ao gerar o arquivo, ele o abre automaticamente.
    # return = Retorna o nome do arquivo para o app (Flask) enviar ao usuário. 

    try:
        abas = pd.read_excel(arquivo_excel, sheet_name=None)
        relatorio_completo = pd.concat(abas.values(), ignore_index=True)
        
        relatorio_final = relatorio_completo.dropna() 

        nome_arquivo_saida = "Relatorio_Completo.xlsx"
        relatorio_final.to_excel(nome_arquivo_saida, index=False)

        os.startfile(nome_arquivo_saida) 

        return nome_arquivo_saida
    
    # Caso dê errado ele irá rodar o except:

    # print = Exibe a mensagem de erro ao gerar relatório no terminal.
    # return = Retorna vazio para o app (Flask).

    except Exception as e:
        print(f"Erro ao gerar relatório completo: {e}")
        return None
        
def gerar_relatorio_funcionarios():
    try:
        abas = pd.read_excel(arquivo_excel, sheet_name=None)
        relatorio_completo = pd.concat(abas.values(), ignore_index=True)
        
        relatorio_final = relatorio_completo.dropna(subset=["Funcionário", "Data do Exame"])
        # relatorio final = Limpa o relatório, retirando as partes em branco (NaN),
        # aqui também selecionamos a coluna que queremos que limpe no relatório.

        relatorio_exibicao = relatorio_final[["Funcionário", "Data do Exame"]]
        # relatorio_exibicao = Definimos as colunas limpas, que iremos exibir no relatório.

        nome_arquivo_saida = "Relatorio_Funcionarios.xlsx" 

        relatorio_exibicao.to_excel(nome_arquivo_saida, index=False)

        os.startfile(nome_arquivo_saida) 

        return nome_arquivo_saida
    except Exception as e:

        print(f"Erro ao gerar relatório de funcionários: {e}")
        return None

def gerar_relatorio_exames():
    try:
        abas = pd.read_excel(arquivo_excel, sheet_name=None)
        relatorio_completo = pd.concat(abas.values(), ignore_index=True)

        relatorio_final = relatorio_completo.dropna(subset=["Exames", "Data do Exame"])

        relatorio_exibicao = relatorio_final[["Exames", "Data do Exame"]]

        nome_arquivo_saida = "Relatorio_Exames.xlsx"

        relatorio_exibicao.to_excel(nome_arquivo_saida, index = False)

        os.startfile(nome_arquivo_saida) 

        return nome_arquivo_saida
    except Exception as e:

        print(f"Erro ao gerar relatório de exames: {e}")
        return None
    
def gerar_relatorio_funcao():
    try:
        abas = pd.read_excel(arquivo_excel, sheet_name=None)
        relatorio_completo = pd.concat(abas.values(), ignore_index=True)

        relatorio_final = relatorio_completo.dropna(subset=["Função", "Data do Exame"])

        relatorio_exibicao = relatorio_final[["Função", "Data do Exame"]]

        nome_arquivo_saida = "Relatorio_Funcao.xlsx"

        relatorio_exibicao.to_excel(nome_arquivo_saida, index = False)

        os.startfile(nome_arquivo_saida) 

        return nome_arquivo_saida
    except Exception as e:

        print(f"Erro ao gerar relatório de função: {e}")
        return None
    
def gerar_graficos_funcionarios():

    try:
        # 1. LEITURA E PREPARAÇÃO DOS DADOS
        abas = pd.read_excel(arquivo_excel, sheet_name=None)
        dados_completos = pd.concat(abas.values(), ignore_index=True)
        
        # Filtragem de dados
        dados_filtrados = dados_completos[
            dados_completos['Funcionário'].notna() & 
            (dados_completos['Funcionário'].astype(str).str.strip() != '') &
            (~dados_completos['Funcionário'].astype(str).str.lower().isin(['funcionário', 'funcionario', '']))
        ]
        
        exames_por_funcionario = dados_filtrados['Funcionário'].value_counts().head(10)
        
        # 2. CONFIGURAÇÃO E GERAÇÃO DO GRÁFICO (USANDO SUBPLOTS)
        
        # MUDANÇA 1: Usando plt.subplots para melhor controle do eixo/área
        fig, ax = plt.subplots(figsize=(12, 5)) 
        
        # Cores (manter o degradê verde)
        cores = plt.cm.Greens(np.linspace(0.4, 0.9, len(exames_por_funcionario)))
        cores[0] = [0.2, 0.6, 0.2, 1]
        
        # Cria barras horizontais usando ax.barh
        bars = ax.barh(exames_por_funcionario.index, exames_por_funcionario.values,
                      color=cores, height=0.7, edgecolor='white', linewidth=1)
        
        # Destaca o primeiro colocado
        bars[0].set_edgecolor('darkgreen')
        bars[0].set_linewidth(2)
        
        # Adiciona os valores nas barras (labels)
        for bar, valor in zip(bars, exames_por_funcionario.values):
            # Usando ax.text
            ax.text(valor + 1, bar.get_y() + bar.get_height()/2, f'{valor}',
                     va='center', fontsize=9, fontweight='bold')
                     
        # Formatação
        ax.set_title('TOP 10 FUNCIONÁRIOS - MAIS EXAMES REALIZADOS\n (2022-2025)', fontsize=12, fontweight='bold', pad=10)
        ax.set_xlabel('Quantidade de Exames', fontsize=10)
        ax.set_ylabel('') # Remove o label Y para economizar espaço
        
        # Reduz o tamanho da fonte dos rótulos do eixo Y
        ax.tick_params(axis='y', labelsize=8) 
        ax.invert_yaxis() 
        
        # Remove as bordas do plot
        for spine in ax.spines.values():
            spine.set_visible(False)
            
        # MUDANÇA 2: Usamos o tight_layout, mas ele será corrigido pelo bbox_inches
        plt.tight_layout() 

        # 3. SALVAR GRÁFICO NO ARQUIVO EXCEL
        
        wb = Workbook()
        img = io.BytesIO()

        # Usando 'fig' (o objeto figure do subplots) e DPI baixo para forçar compactação
        fig.savefig(img, format="png", bbox_inches="tight", dpi=100) 
        
        img.seek(0) 

        planilha_ativa = wb.active
        planilha_ativa.add_image(drawing.image.Image(img), 'A1')

        nome_arquivo_saida = "TOP10_Examinadores.xlsx"
        wb.save(nome_arquivo_saida)

        os.startfile(nome_arquivo_saida)
        
        return nome_arquivo_saida
    
    except Exception as e:
        print(f"Erro ao gerar gráfico de funcionários: {e}")
        return None
    
def gerar_graficos_exames():
    
    try:
        
        # 1. LEITURA E PREPARAÇÃO DOS DADOS (PANDAS)
        # Lê todas as abas do arquivo Excel e concatena os dados.
        abas = pd.read_excel(arquivo_excel, sheet_name=None)
        dados_completos = pd.concat(abas.values(), ignore_index=True)

        # 2. CONFIGURAÇÃO E GERAÇÃO DO GRÁFICO (MATPLOTLIB)
        
        # Define a figura e seu tamanho.
        plt.figure(figsize=(12, 5))
        
        # Calcula o TOP 10 dos exames realizados.
        top_exames = dados_completos['Exames'].value_counts().head(10) 

        # Gera o gráfico de barras horizontal com o mapa de cores 'plasma'.
        plt.barh(top_exames.index, top_exames.values, color=plt.cm.plasma(range(10))) 
        
        # Define título, rótulo do eixo X e inverte o eixo Y.
        plt.title(f'TOP 10 EXAMES REALIZADOS\n (2022-2025)')
        plt.xlabel('Quantidade de Exames')
        plt.gca().invert_yaxis() 
        
        # Ajusta o layout para evitar cortes e pega a figura.
        plt.tight_layout() 
        grafico_completo = plt.gcf()

        # 3. SALVAR GRÁFICO NO ARQUIVO EXCEL (OPENPYXL)
        
        # Inicializa o arquivo Excel e o buffer de imagem.
        wb = Workbook()
        img = io.BytesIO()

        # Salva o gráfico como PNG no buffer e reposiciona o ponteiro.
        grafico_completo.savefig(img, format="png", bbox_inches="tight")
        img.seek(0) 

        # Adiciona a imagem à primeira célula do Excel.
        planilha_ativa = wb.active
        planilha_ativa.add_image(drawing.image.Image(img), 'A1')

        # Salva o arquivo final e o abre.
        nome_arquivo_saida = "Grafico_Exames.xlsx"
        wb.save(nome_arquivo_saida)
        os.startfile(nome_arquivo_saida)
        
        # Retorna o nome do arquivo.
        return nome_arquivo_saida
    
    except Exception as e:
        # Em caso de erro, exibe a mensagem e retorna None.
        print(f"Erro ao gerar gráfico de exames: {e}")
        return None

def gerar_graficos_funcao():
    """Gera gráfico das TOP 15 funções com mais exames"""
    try:
        # Carrega e processa dados
        abas = pd.read_excel("bigData.xlsx", sheet_name=None)
        dados = pd.concat(abas.values(), ignore_index=True)
        
        # Top 15 funções
        top_funcoes = dados['Função'].value_counts().head(15)
        
        # Configura e plota gráfico
        plt.figure(figsize=(12, 6))
        plt.barh(top_funcoes.index, top_funcoes.values, color=plt.cm.viridis(range(15)))
        plt.title(f'TOP 15 FUNÇÕES\n (2022-2025)')
        plt.gca().invert_yaxis()
        
        # Salva no Excel
        wb = Workbook()
        img = io.BytesIO()
        plt.savefig(img, format="png", bbox_inches="tight")
        wb.active.add_image(drawing.image.Image(img), 'A1')
        
        nome_arquivo = "Grafico_Funcao.xlsx"
        wb.save(nome_arquivo)
        os.startfile(nome_arquivo)
        
        return nome_arquivo
        
    except Exception as e:
        print(f"Erro: {e}")
        return None